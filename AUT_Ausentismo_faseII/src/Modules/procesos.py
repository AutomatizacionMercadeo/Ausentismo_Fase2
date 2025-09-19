import holidays
import os
import unicodedata
import re
import datetime as dt
import shutil
from openpyxl import load_workbook
from datetime import datetime, timedelta
from src.Emails.crear_correos import crearCorreos
from src.Emails.error_correo import enviar_error_correo
from src.SQL.consultar_maestra import consultar_maestra_DB
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


class Cruce_datos:
    def __init__(self):
        # self.ruta_justificacion = self.buscar_archivo_justificacion()

        # creamos la ruta de la carpeta reportes
        self.ruta_maestra = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reportes', f'MAESTRAS (Ausentismos).xlsx')
        self.ruta_ausentismo_sin_justificacion = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reportes', f'Ausentismos_SIN_JUSTIFICACION_GENERAL - {self.obtener_ultimo_dia_anterior()}.xlsx')
        #self.ruta_justificacion = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reportes', f'AUSENTISMO_{self.obtener_ultimo_dia_anterior()}.xlsx')
        #self.ruta_ausentismo_vencidos = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reportes', f'Ausentismos_SIN_JUSTIFICACION_GENERAL - {self.fecha_vencida_2_dias}.xlsx')
        self.ruta_reportes_vencidos = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reportes')
        #self.ruta_reportes_consolidado_mensual = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reportes', f"REPORTE_CONSOLIDADO_AUSENTISMO_{self.obtener_nombre_mes(ayer)}_{año_actual}.xlsx")
        self.ruta_reportes_consolidado_diario = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reportes', f"REPORTE_CONSOLIDADO_AUSENTISMO_DIARIO_{self.obtener_ultimo_dia_anterior()}.xlsx")
        self.crear_correos = crearCorreos(correo_destinatario = None)
        print(self.ruta_maestra)

        print(f"Ruta del archivo de justificación: {self.ruta_ausentismo_sin_justificacion}")


    def obtener_ultimo_dia_anterior(self):
        """Obtiene el último día hábil anterior a hoy, evitando domingos y festivos."""
        festivos_co = holidays.Colombia()
        hoy = datetime.now()

        # Si hoy es domingo o festivo, no ejecuta
        if hoy.weekday() == 6 or hoy in festivos_co:
            print("Hoy es domingo o festivo. No se ejecutará el cruce.")
            return None

        dia_anterior = hoy - timedelta(days=1)
        while dia_anterior.weekday() == 6 or dia_anterior in festivos_co:  # 6 = domingo
            dia_anterior -= timedelta(days=1)
        
        return dia_anterior.strftime('%Y-%m-%d')  # Retorna la fecha como string en formato 'YYYY-MM-DD'
        # return "2025-09-04" # para pruebas
    def normalizar_texto(self, texto):
        """Normaliza el texto eliminando espacios en blanco y convirtiendo a minúsculas."""
        if texto is None:
            return ""
        
        # Pasar a minúsculas
        texto = texto.lower()

        # Eliminar acentos/diacríticos
        texto = unicodedata.normalize("NFD", texto)
        texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    
        # Eliminar puntos al final de palabras (ej: 'antiguedad.' -> 'antiguedad')
        texto = re.sub(r"\.+\b", "", texto)


        # Quitar espacios iniciales/finales y reemplazar múltiples espacios internos por 1
        texto = re.sub(r"\s+", " ", texto).strip()
        return texto

    

    def estilo_formato_excel(self, hoja):
        """
        Aplica formato a una hoja de Excel:
        - Encabezado con fondo azul y texto blanco en negrita.
        - Ajuste automático del ancho de columnas.
        - Filtro en la primera fila.
        """
        fill_azul = PatternFill(start_color="1f2a37", end_color="1f2a37", fill_type="solid")
        font_blanco_negrita = Font(color="FFFFFF", bold=True)

        # Encabezados
        for col_num in range(1, hoja.max_column + 1):
            celda = hoja.cell(row=1, column=col_num)
            celda.fill = fill_azul
            celda.font = font_blanco_negrita

        # Ajustar ancho de columnas
        for col in hoja.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            hoja.column_dimensions[col_letter].width = max_length + 2

        # Agregar filtro automático
        hoja.auto_filter.ref = hoja.dimensions


    # ahora creamos una funcion para hacer el cruce entre la maestra y justificacion
    def extraer_datos_JUSTIFICACION(self):
        carpeta_reportes = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reportes")
        fecha = self.obtener_ultimo_dia_anterior()

        patron = re.compile(rf"^REPORTE_AUSENTISMO.*_{fecha}\.xlsx$", re.IGNORECASE)

        archivo_valido = None
        for archivo in os.listdir(carpeta_reportes):
            if patron.match(archivo):
                archivo_valido = os.path.join(carpeta_reportes, archivo)
                print(f"Archivo de justificación encontrado: {archivo_valido}")
                break

        if not archivo_valido:
            print(f"No se encontró un archivo válido de justificación para la fecha {fecha}")
            return None, None

        excel_justificacion = load_workbook(archivo_valido)
        hoja_justificacion = excel_justificacion.active

        datos_justificacion = []
        for fila in hoja_justificacion.iter_rows(min_row=2, values_only=True):
            fila_procesada = []
            for celda in fila:
                if isinstance(celda, datetime):
                    fila_procesada.append(celda.strftime("%d/%m/%Y"))
                else:
                    fila_procesada.append(str(celda) if celda is not None else "")
            datos_justificacion.append(fila_procesada)

        return datos_justificacion, archivo_valido


    # ahora extraemos los datos de la maestra
    def extraer_datos_maestra(self):
        # Lógica para extraer datos de la tabla maestra

        if not os.path.exists(self.ruta_maestra):
            print(f"El archivo {self.ruta_maestra} no existe.")
            return False

        # cargamos el archivo maestra excel
        excel_maestra = load_workbook(self.ruta_maestra)
        
        #seleccionamos la hoja permisos y beneficios
        hoja_maestra = excel_maestra['Permisos y Beneficios']  # HOJA Permisos y Beneficios

        # ahora extraemos la columna Nombre del Beneficio
        nombres_beneficio = []
        for fila in hoja_maestra.iter_rows(min_row=2, values_only=True):
            nombre = fila[1]  # el nombre del beneficio está en la segunda columna
            nombres_beneficio.append(nombre)

        # print("Nombres de los beneficios extraídos:")
        # for nombre in nombres_beneficio:
        #     print(nombre)

        return nombres_beneficio
    
    def extraer_datos_ausentismo_sin_justificacion(self):
        # Lógica para extraer datos de la tabla Ausentismo sin Justificación
        if not os.path.exists(self.ruta_ausentismo_sin_justificacion):
            print(f"El archivo {self.ruta_ausentismo_sin_justificacion} no existe.")
            return False

        # cargamos el archivo ausentismo sin justificacion excel
        excel_ausentismo = load_workbook(self.ruta_ausentismo_sin_justificacion)
        # seleccionamos la hoja activa
        hoja_ausentismo = excel_ausentismo.active

        # extraemos datos de la hoja
        datos_ausentismo = []
        for fila in hoja_ausentismo.iter_rows(min_row=2, values_only=True):
            fila_procesada = []
            for celda in fila:
                if isinstance(celda, datetime):  
                    fila_procesada.append(celda.strftime("%Y-%m-%d"))
                else:
                    fila_procesada.append(str(celda) if celda is not None else "")
            datos_ausentismo.append(fila_procesada)

    
        # print("Datos extraídos de ausentismo sin justificación:")
        # for fila in datos_ausentismo:
        #     print(fila)

        return datos_ausentismo

    # ahora creamos una funcion para hacer el cruce entre los dos archivos
    def cruce_datos_ausencia_vs_beneficio(self, dia_anterior):
        datos_justificacion, ruta_justificacion = self.extraer_datos_JUSTIFICACION()
        nombres_beneficio = self.extraer_datos_maestra()

        print("Datos de justificación extraídos:", datos_justificacion)

        if not datos_justificacion or not nombres_beneficio:
            print("No se pudieron extraer los datos necesarios para el cruce.")
            return
        
        data_procesada = []
        list_no_coincidencia = []

        print("Datos cruzados:")
        for fila_justificacion in datos_justificacion:
            if not any(fila_justificacion):  # fila vacía
                continue

            valor_justificacion = self.normalizar_texto(fila_justificacion[5])  
            coincidencia = False

            for nombre_beneficio in nombres_beneficio:
                valor_beneficio = self.normalizar_texto(nombre_beneficio)

                if valor_justificacion == valor_beneficio:
                    
                    # Validación de fechas
                    fecha_inicial_str = fila_justificacion[7]  # Ajusta índice según Excel
                    fecha_final_str   = fila_justificacion[8]  # Ajusta índice según Excel

                    try:
                        fecha_inicial = datetime.strptime(fecha_inicial_str, "%d/%m/%Y")
                        fecha_final   = datetime.strptime(fecha_final_str, "%d/%m/%Y")
                    except Exception:
                        print(f"Error de formato en fechas: {fecha_inicial_str}, {fecha_final_str}")
                        list_no_coincidencia.append(fila_justificacion)
                        coincidencia = True  # Para que no entre a no coincidencia otra vez
                        break

                    # Año actual
                    anio_actual = datetime.now().year
                    if fecha_inicial.year != anio_actual or fecha_final.year != anio_actual:
                        print(f"Error: Fechas no corresponden al año actual. {fecha_inicial_str} - {fecha_final_str}")
                        list_no_coincidencia.append(fila_justificacion)
                        coincidencia = True
                        break

                    # Validar rango
                    if fecha_final < fecha_inicial:
                        print(f"Error: La fecha final {fecha_final_str} es menor que la inicial {fecha_inicial_str}")
                        list_no_coincidencia.append(fila_justificacion)
                        coincidencia = True
                        break

                
                    # Si pasa todas las validaciones
                    print(f"Coincidencia encontrada: {fila_justificacion} - {nombre_beneficio}")
                    data_procesada.append(fila_justificacion)
                    coincidencia = True
                    break  

            if not coincidencia:
                print(f"No hay coincidencia: {fila_justificacion}")
                list_no_coincidencia.append(fila_justificacion)

        print(f"Total de no coincidencias encontradas: {len(list_no_coincidencia)}")

        if list_no_coincidencia:
            wb = load_workbook(ruta_justificacion)
            ws = wb.active

            ws.delete_rows(2, ws.max_row)  # limpiar filas

            for fila in list_no_coincidencia:
                ws.append(fila)

            wb.save(ruta_justificacion)

            data = consultar_maestra_DB()
            if not data:
                print("No se encontraron datos en la maestra.")
                return False

            errores_por_centro = {}
            correos_fijos = "jose.chaverra@gruporeditos.com"
            for fila_error in list_no_coincidencia:
                centro_de_costo = fila_error[1]  
                valor_justificacion = fila_error[5]

                correo_destinatario = None
                for fila_maestra in data:
                    if fila_maestra.get('CENTRO_COSTOS') == centro_de_costo:
                        correo_destinatario = fila_maestra.get('CORREO')
                        print(f"Correo encontrado: {correo_destinatario}")
                        break

                if not correo_destinatario:
                    print(f"No se encontró correo para el centro de costo: {centro_de_costo}")
                    continue

                if correo_destinatario not in errores_por_centro:
                    errores_por_centro[correo_destinatario] = []
                    
                errores_por_centro[correo_destinatario].append(
                    f"Motivo de ausencia: '{valor_justificacion}'\n"
                    f"Usuario detectado: {fila_error[4]}\n"
                    f"Cédula: {fila_error[3]}\n"
                    f"Fechas: {fila_error[7]} - {fila_error[8]}\n"
                )

            for correo_destinatario, errores in errores_por_centro.items():
                message = (
                    "Error: Se encontraron inconsistencias en las fechas o motivos de ausencia:\n\n"
                    + "\n".join(errores) +
                    "\nPor favor revise la información y reenvíe el archivo de justificación corregido."
                )
                destinatarios = [correo_destinatario, correos_fijos]
                enviar_error_correo(message, ruta_justificacion, dia_anterior, destinatarios)
        else:
            print("No se encontraron discrepancias para actualizar en justificacion.")
            
        os.remove(ruta_justificacion)
        return data_procesada





    # ESTA FUNCION ES PARA CRUZAR LOS DATOS DE CEDULA DE JUSTIFICACION CON CEDULA REPORTE AUSENTISMO_SIN_JUSITIFACION DE LA DFASE I
    def cruce_datos_cedula_vs_cedula(self, dia_anterior):


        """
        Cruza los datos de cédulas de justificación vs ausentismo sin justificación.
        Si hay coincidencia en la cédula:
            - Se guarda en data_consolidada (para consolidado general).
            - Se elimina de datos_ausentismo (no debe seguir como pendiente).
        """
        datos_ausentismo = self.extraer_datos_ausentismo_sin_justificacion()
        data_procesada = self.cruce_datos_ausencia_vs_beneficio(dia_anterior)

        if not datos_ausentismo or not data_procesada:
            print("No se pudieron extraer los datos necesarios para el cruce de cedulas vs cedulas.")
            return []

        data_consolidada = []       # Coincidencias -> van al consolidado
        ausentismo_restante = []    # Sin coincidencias -> quedan en SIN JUSTIFICACION

        # Normalizamos las cédulas al comparar (evita int/str/espacios)
        for fila_ausentismo in datos_ausentismo:
            cedula_ausente = str(fila_ausentismo[3]).strip()
            coincidencia = False

            for fila_justificacion in data_procesada:
                cedula_just = str(fila_justificacion[3]).strip()
                if cedula_ausente == cedula_just:
                    print(f"Coincidencia encontrada: {fila_ausentismo} - {fila_justificacion}")
                    data_consolidada.append(fila_justificacion)   # Guardamos en consolidado
                    coincidencia = True
                    break

            if not coincidencia:
                ausentismo_restante.append(fila_ausentismo)  # Solo guardamos los que NO tienen justificación

        # ---------- Actualizar el archivo SIN JUSTIFICACION: siempre escribimos (incluso si queda vacío) ----------
        try:
            wb = load_workbook(self.ruta_ausentismo_sin_justificacion)
            ws = wb.active

            # Borrar todas las filas de datos (dejamos el encabezado en la fila 1 si existe)
            if ws.max_row > 1:
                # amount = número de filas a borrar empezando desde la fila 2
                amount = ws.max_row - 1
                ws.delete_rows(2, amount)

            # Escribir los registros que quedaron (puede ser una lista vacía)
            for fila in ausentismo_restante:
                ws.append(fila)

            wb.save(self.ruta_ausentismo_sin_justificacion)
            print(f"Actualizando Ausentismos_SIN_JUSTIFICACION en {self.ruta_ausentismo_sin_justificacion}")

        except Exception as e:
            print(f"Error actualizando {self.ruta_ausentismo_sin_justificacion}: {e}")

        return data_consolidada

    

    def reporte_consolidado_diario(self, data_consolidada):
        
        if not os.path.exists(self.ruta_reportes_consolidado_diario):
            workbook = Workbook()
            hoja = workbook.active
            hoja.title = "Consolidado"

            # Escribimos el encabezado SOLO si es un archivo nuevo
            headers = [
                'ZONA', 'CENTRO_COSTOS', 'OFICINA', 'CEDULA', 'NOMBRE', 
                'MOTIVO_AUSENCIA', 'DIAS', 'FECHA_INICIAL', 'FECHA_FINAL'
            ]
            hoja.append(headers)

            # Escribir los datos procesados (nuevos)
            for fila in data_consolidada:
                hoja.append(fila)

            # Aplicamos formato SOLO si es nuevo (para no repetir estilo cada vez)
            self.estilo_formato_excel(hoja)

            # Guardar el archivo
            workbook.save(self.ruta_reportes_consolidado_diario)
            print(f"Reporte consolidado actualizado en: {self.ruta_reportes_consolidado_diario}")

            













# if __name__ == "__main__":
#     cruce = Cruce_datos()
#     dia_anterior = cruce.obtener_ultimo_dia_anterior()


#     # cruce.extraer_datos_JUSTIFICACION()
#     # cruce.extraer_datos_maestra()
#     cruce.cruce_datos_ausencia_vs_beneficio(dia_anterior)
#     # cruce.extraer_datos_ausentismo_sin_justificacion()
#     # cruce.cruce_datos_cedula_vs_cedula(dia_anterior)