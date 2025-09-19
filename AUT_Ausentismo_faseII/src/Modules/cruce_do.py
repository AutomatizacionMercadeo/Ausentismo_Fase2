import os
import re
import unicodedata
from src.SQL.consultar_consolidado_mensual import consultar_consolidado_mensual
from datetime import datetime
from openpyxl import Workbook, load_workbook
from src.Modules.procesos import Cruce_datos    


def normalizar_texto(texto):
        """Normaliza el texto eliminando espacios en blanco y convirtiendo a minúsculas."""
        if texto is None:
            return ""
        
        # Pasar a minúsculas
        texto = texto.lower()

        # Eliminar acentos/diacríticos
        texto = unicodedata.normalize("NFD", texto)
        texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    
        # Eliminar puntos al final de palabras (ej: 'antiguedad.' -> 'antiguedad')
        texto = re.sub(r"\.+\b", "", texto)


        # Quitar espacios iniciales/finales y reemplazar múltiples espacios internos por 1
        texto = re.sub(r"\s+", " ", texto).strip()
        return texto



def extraer_reporte_do():

    # obtener la ruta src
    path_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ruta_reporte_DO = os.path.join(path_root, 'Reportes_Ausentismos', 'Reporte_DO.xlsx')

    print(path_root)
    print(ruta_reporte_DO)

    # validamos si la ruta existe
    try:
        if os.path.exists(ruta_reporte_DO):
            print(f"El archivo {ruta_reporte_DO} existe.")
        else:
            print(f"El archivo {ruta_reporte_DO} no existe.")
            return False
    except Exception as e:
        print(f"Ocurrió un error al verificar el archivo: {e}")
        return False
    

    excel_do = load_workbook(ruta_reporte_DO)
    hoja_do = excel_do.active

    # extraemos datos de la hoja
    datos_do = []
    for fila in hoja_do.iter_rows(min_row=2, values_only=True):
        fila_procesada = []
        for celda in fila:
            if isinstance(celda, datetime):  
                fila_procesada.append(celda.strftime("%d/%m/%Y"))
            else:
                fila_procesada.append(str(celda) if celda is not None else "")
        datos_do.append(fila_procesada)

    return datos_do

# ahora creamos una funcion para guardar los datos no coincidentes en un archivo excel
def guardar_datos_no_coincidentes(datos_no_coincidentes, dia_habil):

    if not datos_no_coincidentes:
        print("No hay datos no coincidentes para guardar.")
        return

    # obtener la ruta src
    path_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ruta_guardado = os.path.join(path_root, 'Reportes_Ausentismos', f'Datos_No_Coincidentes_con_DO_{dia_habil}.xlsx')

    workbook = load_workbook(ruta_guardado) if os.path.exists(ruta_guardado) else Workbook()
    hoja = workbook.active
    hoja.title = "No Coincidentes"

    # Definimos los encabezados
    headers = ['ZONA', 'CENTRO_COSTOS', 'OFICINA', 'CEDULA', 'NOMBRE', 
                'MOTIVO_AUSENCIA', 'DIAS', 'FECHA_INICIAL', 'FECHA_FINAL']

    # Reemplazamos siempre la primera fila con los encabezados
    for col, header in enumerate(headers, start=1):
        hoja.cell(row=1, column=col, value=header)

    # Agregamos los datos a partir de la fila 2
    for fila in datos_no_coincidentes:
        hoja.append([
            fila.get('ZONA', ''),
            fila.get('CENTRO_COSTOS', ''),
            fila.get('OFICINA', ''),
            fila.get('CEDULA', ''),
            fila.get('NOMBRE', ''),
            fila.get('MOTIVO_AUSENCIA', ''),
            fila.get('DIAS', ''),
            fila.get('FECHA_INICIAL', ''),
            fila.get('FECHA_FINAL', '')
        ])

    # llamamos a la funcion de estilo para dar formato a la hoja
    estilo = Cruce_datos()
    estilo.estilo_formato_excel(hoja)

    try:
        workbook.save(ruta_guardado)
        print(f"Datos no coincidentes guardados en {ruta_guardado}")
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")


def cruce_do(dia_habil):
    try:
        datos_do = extraer_reporte_do()
        datos_consolidado = consultar_consolidado_mensual()

        if not datos_do or not datos_consolidado:
            print("No se pudieron extraer los datos necesarios.")
            return

        # cruzamos los datos por cedula
        
        datos_no_coincidentes = []
        for fila_consolidado in datos_consolidado:
            for fila_do in datos_do:
                if fila_do[3] == fila_consolidado["CEDULA"]: # validamos las cedulas

                    # normalizamos los motivos de ausencia para hacer la validacion
                    motivo_ausencia_do = normalizar_texto(fila_do[5])
                    motivo_ausencia_consolidado = normalizar_texto(fila_consolidado["MOTIVO_AUSENCIA"])

                    if motivo_ausencia_do in motivo_ausencia_consolidado: # validamos el motivo de ausencia
                        # ahora validamos las fechas
                    
                        fecha_inicial_do = fila_do[7]
                        fecha_final_do = fila_do[8]

                        fecha_inicial_consolidado = fila_consolidado["FECHA_INICIAL"]
                        fecha_final_consolidado = fila_consolidado["FECHA_FINAL"]

                        if fecha_inicial_do == fecha_inicial_consolidado and fecha_final_do == fecha_final_consolidado:
                            # si todo coincide, agregamos la fila combinada a datos_cruzados
                            print(f"Coincidencia encontrada para CEDULA {fila_do[3]}")
                            break  # salimos del loop interno para la siguiente fila_do
                        else:
                            print(f"Fechas no coinciden para CEDULA {fila_consolidado['CEDULA']}: DO({fecha_inicial_do} - {fecha_final_do}) vs Consolidado({fecha_inicial_consolidado} - {fecha_final_consolidado})")
                            datos_no_coincidentes.append(fila_consolidado)
                            break
                    else:
                        print(f"Motivo de ausencia no coincide para CEDULA {fila_consolidado['CEDULA']}: '{fila_do[5]}' vs '{fila_consolidado['MOTIVO_AUSENCIA']}'")
                        datos_no_coincidentes.append(fila_consolidado)
                        break
                else:
                    print(f"No se encontró CEDULA {fila_consolidado['CEDULA']} en el consolidado.")
                    datos_no_coincidentes.append(fila_consolidado)
                    break


        guardar_datos_no_coincidentes(datos_no_coincidentes, dia_habil)
        return True

    except Exception as e:
        print(f"Error durante el cruce de datos: {e}")
        return False

