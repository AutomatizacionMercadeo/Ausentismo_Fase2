import os
from openpyxl import Workbook, load_workbook
import datetime as dt
from datetime import datetime, timedelta
from src.Modules.procesos import Cruce_datos
# Diccionario con los meses en español
MESES_DIC = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }


def crear_carpeta_consolidado_mensual():
    """Crea la carpeta 'Consolidado_Mensual' si no existe."""
    folder = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Consolidado_Mensual")
    if not os.path.exists(folder):
        os.mkdir(folder)
        print(f"Carpeta 'Consolidado_Mensual' creada en: {folder}")
    else:
        print(f"La carpeta 'Consolidado_Mensual' ya existe en: {folder}")


def obtener_nombre_mes(fecha: datetime) -> str:
    """Devuelve el nombre del mes en español en mayúsculas."""
    return MESES_DIC[fecha.month]


año_actual = datetime.now().year
ayer = datetime.now() - timedelta(days=1)
ruta_reportes_consolidado_mensual = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Consolidado_Mensual', f"REPORTE_CONSOLIDADO_AUSENTISMO_{obtener_nombre_mes(ayer)}_{año_actual}.xlsx")



def extraer_datos_consolidado_diario(dia_anterior):
    ruta_reportes_consolidado_diario = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Reportes_Ausentismos', f"REPORTE_CONSOLIDADO_AUSENTISMO_DIARIO_{dia_anterior}.xlsx")
    if not os.path.exists(ruta_reportes_consolidado_diario):
        print(f"No se encontró el archivo diario en: {ruta_reportes_consolidado_diario}")
        return False
    try:
        excel_consolidado_diario = load_workbook(ruta_reportes_consolidado_diario)
        hoja_consolidado = excel_consolidado_diario.active


        datos_consolidado_diario = []
        for fila in hoja_consolidado.iter_rows(min_row=2, values_only=True):  # Saltar la fila de encabezado
            datos_consolidado_diario.append(list(fila))
        return datos_consolidado_diario

    except Exception as e:
        print(f"Error al leer el archivo diario: {e}")
        return False





def reporte_consolidado_mensual(dia_anterior):

    # creamos la caerpeta si no existe
    crear_carpeta_consolidado_mensual()

    #obtener los datos del consolidado diario
    data_consolidada = extraer_datos_consolidado_diario(dia_anterior)

    if os.path.exists(ruta_reportes_consolidado_mensual):
        workbook = load_workbook(ruta_reportes_consolidado_mensual)
        hoja = workbook.active
        start_row = hoja.max_row + 1
    else:
        workbook = Workbook()
        hoja = workbook.active
        hoja.title = "Consolidado"
        start_row = 1

        headers = [
            'ZONA', 'CENTRO_COSTOS', 'OFICINA', 'CEDULA', 'NOMBRE', 
            'MOTIVO_AUSENCIA', 'DIAS', 'FECHA_INICIAL', 'FECHA_FINAL'
        ]
        hoja.append(headers)

    if data_consolidada:
        for fila in data_consolidada:
            cedula_nueva = fila[3]
            fecha_inicial = fila[7]
            fecha_final = fila[8]

            repetido = False
            for fila_existente in hoja.iter_rows(min_row=2, values_only=True):
                if (cedula_nueva == fila_existente[3] and 
                    fecha_inicial == fila_existente[7] and 
                    fecha_final == fila_existente[8]):
                    repetido = True
                    print(f"Registro ya existe y no se añadirá: CEDULA {cedula_nueva}, FECHA INICIAL {fecha_inicial}, FECHA FINAL {fecha_final}")
                    break

            if not repetido:
                hoja.append(fila)
                print(f"Registro añadido: CEDULA {cedula_nueva}, FECHA INICIAL {fecha_inicial}, FECHA FINAL {fecha_final}")

    #  aplicar siempre 
    # if start_row == 1:
    cruce = Cruce_datos()
    cruce.estilo_formato_excel(hoja)

    # Guardar el archivo
    workbook.save(ruta_reportes_consolidado_mensual)
    print(f"Reporte consolidado actualizado en: {ruta_reportes_consolidado_mensual}")






# if __name__ == "__main__":
#     crear_carpeta_consolidado_mensual()
#     print(ruta_reportes_consolidado_mensual)